from __future__ import annotations

import csv
import io
from collections import Counter
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import app
from src.settings import EXPORTS_DIR


def test_quick_mode_end_to_end_generates_table_and_exports() -> None:
    client = TestClient(app)

    run_resp = client.post("/runs", json={"mode": "quick", "wait": True})
    assert run_resp.status_code == 200

    job = run_resp.json()["job"]
    assert job["status"] == "COMPLETED"

    table_resp = client.get("/results/table", params={"job_id": job["id"]})
    assert table_resp.status_code == 200

    payload = table_resp.json()
    assert payload["rows"]

    has_citation_with_confidence = False
    for row in payload["rows"]:
        for cell in row["cells"]:
            if cell.get("citation") and (cell.get("confidence") or 0) > 0:
                has_citation_with_confidence = True
                break
        if has_citation_with_confidence:
            break

    assert has_citation_with_confidence

    before_csv = {p.name for p in EXPORTS_DIR.glob("*.csv")}
    before_xlsx = {p.name for p in EXPORTS_DIR.glob("*.xlsx")}

    csv_resp = client.post("/exports/csv", json={"job_id": job["id"]})
    xlsx_resp = client.post("/exports/xlsx", json={"job_id": job["id"]})

    assert csv_resp.status_code == 200
    assert xlsx_resp.status_code == 200

    after_csv = {p.name for p in EXPORTS_DIR.glob("*.csv")}
    after_xlsx = {p.name for p in EXPORTS_DIR.glob("*.xlsx")}

    assert after_csv - before_csv
    assert after_xlsx - before_xlsx


def test_export_reflects_manual_and_rejected_states() -> None:
    client = TestClient(app)

    run_resp = client.post("/runs", json={"mode": "quick", "wait": True})
    assert run_resp.status_code == 200
    job = run_resp.json()["job"]

    table_resp = client.get("/results/table", params={"job_id": job["id"]})
    assert table_resp.status_code == 200
    payload = table_resp.json()

    cells = [cell for row in payload["rows"] for cell in row["cells"] if cell.get("cell_id")]
    assert len(cells) >= 2

    manual_cell = cells[0]
    reject_cell = cells[1]

    manual_resp = client.patch(
        f"/cells/{manual_cell['cell_id']}",
        json={"actor": "test", "manual_value": "manual override value", "reason": "regression"},
    )
    assert manual_resp.status_code == 200
    assert manual_resp.json()["cell"]["review_state"] == "MANUAL_UPDATED"

    reject_resp = client.patch(
        f"/cells/{reject_cell['cell_id']}",
        json={"actor": "test", "review_state": "REJECTED", "reason": "regression"},
    )
    assert reject_resp.status_code == 200
    assert reject_resp.json()["cell"]["review_state"] == "REJECTED"

    csv_resp = client.post("/exports/csv", json={"job_id": job["id"]})
    assert csv_resp.status_code == 200

    csv_text = csv_resp.content.decode("utf-8")
    rows = list(csv.DictReader(io.StringIO(csv_text)))
    states = {row["review_state"] for row in rows}

    assert "MANUAL_UPDATED" in states
    assert "REJECTED" in states


def test_xlsx_matches_csv_for_same_job_and_reviews() -> None:
    client = TestClient(app)

    run_resp = client.post("/runs", json={"mode": "quick", "wait": True})
    assert run_resp.status_code == 200
    job = run_resp.json()["job"]
    job_id = job["id"]

    table_resp = client.get("/results/table", params={"job_id": job_id})
    assert table_resp.status_code == 200
    payload = table_resp.json()

    cells = [cell for row in payload["rows"] for cell in row["cells"] if cell.get("cell_id")]
    assert len(cells) >= 2

    manual_resp = client.patch(
        f"/cells/{cells[0]['cell_id']}",
        json={"actor": "test", "manual_value": "manual parity value", "reason": "xlsx parity"},
    )
    reject_resp = client.patch(
        f"/cells/{cells[1]['cell_id']}",
        json={"actor": "test", "review_state": "REJECTED", "reason": "xlsx parity"},
    )
    assert manual_resp.status_code == 200
    assert reject_resp.status_code == 200
    assert manual_resp.json()["cell"]["review_state"] == "MANUAL_UPDATED"
    assert reject_resp.json()["cell"]["review_state"] == "REJECTED"

    csv_resp = client.post("/exports/csv", json={"job_id": job_id})
    xlsx_resp = client.post("/exports/xlsx", json={"job_id": job_id})
    assert csv_resp.status_code == 200
    assert xlsx_resp.status_code == 200

    csv_rows = list(csv.DictReader(io.StringIO(csv_resp.content.decode("utf-8"))))

    workbook = load_workbook(BytesIO(xlsx_resp.content), data_only=True)
    sheet = workbook.active
    header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    xlsx_rows = [dict(zip(header, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

    assert {row["field_key"] for row in csv_rows} == {row["field_key"] for row in xlsx_rows}
    assert Counter(row["review_state"] for row in csv_rows) == Counter(row["review_state"] for row in xlsx_rows)

    def _norm(value: object) -> str:
        if value is None:
            return ""
        return str(value)

    csv_index = {(row["field_key"], row["document_identifier"]): row for row in csv_rows}
    xlsx_index = {(row["field_key"], row["document_identifier"]): row for row in xlsx_rows}
    assert set(csv_index.keys()) == set(xlsx_index.keys())

    compare_cols = [
        "value",
        "value_raw",
        "value_normalized",
        "review_state",
        "citation_location",
        "citation_location_type",
        "citation_char_start",
        "citation_char_end",
        "citation_snippet",
    ]
    for key in csv_index:
        csv_row = csv_index[key]
        xlsx_row = xlsx_index[key]
        for col in compare_cols:
            assert _norm(csv_row.get(col)) == _norm(xlsx_row.get(col))


def test_results_payload_does_not_expose_template_status_text() -> None:
    client = TestClient(app)

    run_resp = client.post("/runs", json={"mode": "quick", "wait": True})
    assert run_resp.status_code == 200
    job_id = run_resp.json()["job"]["id"]

    payload_resp = client.get("/results/table", params={"job_id": job_id})
    assert payload_resp.status_code == 200
    payload = payload_resp.json()

    assert "template" not in payload
    serialized = str(payload)
    assert "Template status" not in serialized
    assert "final_v1_ready" not in serialized
    assert "example_only_pending_user_confirmation" not in serialized
